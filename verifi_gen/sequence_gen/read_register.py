
###read register

import openpyxl, pprint

wb=openpyxl.load_workbook(r"spi_reg_model.xlsx")
sheetnames=wb.sheetnames
sheet_len=len(sheetnames)

regbitdata = {}

##读取每个寄存器的bit位信息:
for i in range(1,sheet_len):
    for row in range(2,wb[sheetnames[i]].max_row+1):
        bit_name   = wb[sheetnames[i]].cell(row,1).value
        bit_site   = wb[sheetnames[i]].cell(row,2).value
        bit_width  = wb[sheetnames[i]].cell(row,3).value
        bit_access = wb[sheetnames[i]].cell(row,4).value
        regbitdata.setdefault(sheetnames[i],{})
        regbitdata[sheetnames[i]].setdefault(bit_name,{
            'bit_site':bit_site,
            'bit_width':bit_width,
            'bit_access':bit_access
        })
# pprint.pprint(regbitdata)

