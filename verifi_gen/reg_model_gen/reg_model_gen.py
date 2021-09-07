#! python3
# generate uvm_reg ralf files

import openpyxl, pprint, os
import sys
import shutil

print('Opening workbook...')
wb = openpyxl.load_workbook(r'../read_gister/can_reg_model.xlsx')

sheet_name = wb.sheetnames
sheet_len = len(sheet_name)
# add absolute path
absolute_path = 'test_top.dut'

regData = {}
regbitData = {}


def reg_bit_read(i):
    print('Reading reg bits row...')
    for row in range(2, wb[sheet_name[i]].max_row + 1):
        # Each row in the spreadsheet has data: reg_name,address,reset_value.
        bit_name = wb[sheet_name[i]]['A' + str(row)].value
        bit_site = wb[sheet_name[i]]['B' + str(row)].value
        bit_width = wb[sheet_name[i]]['C' + str(row)].value
        bit_access = wb[sheet_name[i]]['D' + str(row)].value
        regbitData.setdefault(sheet_name[i], {})
        regbitData[sheet_name[i]].setdefault(bit_name, {'bit_site': bit_site,
                                                        'bit_width': bit_width,
                                                        'bit_access': bit_access})


def reg_ralf_write(j):
    bin_result = []

    ralf.write('register ' + sheet_name[j] + ' {\n')
    rst = regData[sheet_name[2]]['reg_rst']
    # reg_rst bin data
    bin_rst = (bin(int(rst, 16))).lstrip('0b').rjust(32, '0')

    for bit in bin_rst:
        bin_result.append(bit)

    for k, v in regbitData[sheet_name[j]].items():
        ralf.write('  field ')
        ralf.write(k + ' (' + absolute_path + '.' + sheet_name[j] + ') '
                   + '@\'' + str(v['bit_site']) + ' {\n')
        ralf.write('    bits ' + str(v['bit_width']) + ';\n')
        ralf.write('    access ' + str(v['bit_access']) + ';\n')
        ralf.write('    reset ' + '\'h' + bin_result[v['bit_site']]
                   + ';\n')
        ralf.write('  }\n')

    ralf.write('}\n')


def reg_block_write():
    ralf.write('block' + ' reg_model ' + '{\n')
    ralf.write('  bytes 4;\n')
    for key, val in regData.items():
        ralf.write('  register ' + key + '        ' + '(' + key + ')'
                   + '        ' + '@\'h' + str(val['reg_addr']) + '\n')

    ralf.write('}\n')


# main
# read reg_model data
print('Reading row...')
for row in range(2, wb[sheet_name[0]].max_row + 1):
    # Each row in the spreadsheet has data: reg_name,address,reset_value.
    reg_name = wb[sheet_name[0]]['A' + str(row)].value
    reg_addr = wb[sheet_name[0]]['B' + str(row)].value
    reg_rst = wb[sheet_name[0]]['C' + str(row)].value

    regData.setdefault(reg_name, {'reg_addr': reg_addr,
                                  'reg_rst': reg_rst})
pprint.pprint(regData)

# read reg bits data
for inx in range(1, sheet_len):
    reg_bit_read(i=inx)

pprint.pprint(regbitData)

# create reg_model directory
regmodel_dir = 'reg_model'
if os.path.exists(regmodel_dir):
    shutil.rmtree(regmodel_dir)
os.mkdir(regmodel_dir)
print('RESULT DIR: ' + regmodel_dir)
print('')

# Open a new text file and write the contents of reg data to it.
print('Writing ralf...')
# ralf = open('reg.ralf', 'w')
ralf = open(regmodel_dir + os.sep + 'reg.ralf', 'w')
# reg register write
for ins in range(1, sheet_len):
    reg_ralf_write(j=ins)
# reg block write
reg_block_write()

ralf.close()
print('Done.')